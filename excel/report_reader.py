"""
일일보고 Excel에서 오늘 날짜 열을 찾아 보고 데이터를 읽어옴.

Excel 구조 ('일일 보고' 시트):
- 1행: 날짜 헤더 (datetime 객체)
- A~C열: 카테고리 레이블
- 데이터는 날짜에 해당하는 열에 위치
"""

from datetime import date
from pathlib import Path

import openpyxl

from config import DAILY_REPORT_EXCEL_PATH

SHEET_NAME = "일일 보고"

# 행 번호 매핑 (1-based, 실제 파일 구조 기준)
ROW_MAP = {
    # T-VIEW 전체
    "전체_총":       12,
    "전체_2025이후": 16,
    "전체_전일대비": 17,
    # SKS
    "SKS_총":        18,
    "SKS_2025이후":  26,
    "SKS_전일대비":  27,
    # SKB
    "SKB_총":        28,
    "SKB_2025이후":  36,
    "SKB_전일대비":  37,
    # S1
    "S1_총":         38,
    "S1_2025이후":   42,
    "S1_전일대비":   43,
    # SKT
    "SKT_총":        44,
    "SKT_2024이후":  48,
    "SKT_전일대비":  49,
    # T-View 기타
    "기타_총":       50,
    "기타_2025이후": 54,
    "기타_전일대비": 55,
    # 트래픽
    "rx_max":        62,
    "rx_delta":      63,
    "tx_max":        64,
    "tx_delta":      65,
}


def read_daily_report(target_date: date = None) -> dict:
    """target_date 열의 데이터를 읽어 report_formatter가 쓰는 구조로 반환."""
    target_date = target_date or date.today()
    wb = openpyxl.load_workbook(DAILY_REPORT_EXCEL_PATH, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"시트 '{SHEET_NAME}'를 찾을 수 없습니다. 시트 목록: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    col = _find_date_column(ws, target_date)
    if col is None:
        raise ValueError(f"{target_date}에 해당하는 열을 찾을 수 없습니다.")

    def v(key):
        row = ROW_MAP.get(key)
        if row is None:
            return 0
        val = ws.cell(row=row, column=col).value
        return val if val is not None else 0

    return {
        "subscriber": {
            "전체": {"총": v("전체_총"), "누적": v("전체_2025이후"), "전일대비": v("전체_전일대비")},
            "SKS":  {"총": v("SKS_총"),  "누적": v("SKS_2025이후"),  "전일대비": v("SKS_전일대비")},
            "SKB":  {"총": v("SKB_총"),  "누적": v("SKB_2025이후"),  "전일대비": v("SKB_전일대비")},
            "S1":   {"총": v("S1_총"),   "누적": v("S1_2025이후"),   "전일대비": v("S1_전일대비")},
            "SKT":  {"총": v("SKT_총"),  "누적": v("SKT_2024이후"),  "전일대비": v("SKT_전일대비")},
            "기타": {"총": v("기타_총"), "누적": v("기타_2025이후"), "전일대비": v("기타_전일대비")},
        },
        "traffic": {
            "rx_gbps":   v("rx_max"),
            "rx_delta":  v("rx_delta"),
            "tx_gbps":   v("tx_max"),
            "tx_delta":  v("tx_delta"),
            "base_date": target_date.strftime("%Y/%m/%d"),
        },
    }


def _find_date_column(ws, target_date: date) -> int | None:
    """1행에서 target_date와 일치하는 열 번호 반환 (datetime 객체로 저장돼 있음)."""
    for cell in ws[1]:
        if cell.value is None:
            continue
        # datetime 또는 date 객체 모두 처리
        cell_date = cell.value.date() if hasattr(cell.value, "date") else cell.value
        if cell_date == target_date:
            return cell.column
    return None
