"""
장애 데이터를 연간 Excel 파일의 월별 시트에 기록.

파일 구조:
  - 파일명: tview1.0_장애처리현황(YYYY).xlsx
  - 시트명: "YYYY년 M월"
  - 1행: 장애 신고 현황 / 건수
  - 2행: 컬럼 헤더
  - 3행~: 데이터
"""

from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import INCIDENT_EXCEL_DIR

# 컬럼 순서 (실제 파일 2행 헤더 기준)
COLUMNS = [
    "ID", "접수자명", "접수일시", "발생일시", "처리완료일시",
    "고객사", "장애유형", "서버유형", "장애구분", "이슈여부",
    "장애등급", "제목", "내용", "조치내역", "장애원인사유",
    "향후대책", "장애처리의견", "건수", "처리시간(분)"
]

_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 컬럼 인덱스 (1-based)
_COL = {col: idx for idx, col in enumerate(COLUMNS, start=1)}


def get_yearly_filepath(dt: datetime = None) -> Path:
    dt = dt or datetime.now()
    return Path(INCIDENT_EXCEL_DIR) / f"tview1.0_장애처리현황({dt.year}).xlsx"


def get_sheet_name(dt: datetime = None) -> str:
    dt = dt or datetime.now()
    return f"{dt.year}년 {dt.month}월"


def append_incident(record: dict, dt: datetime = None) -> Path:
    """단건 장애 레코드를 해당 월 시트에 추가."""
    dt = dt or datetime.now()
    filepath = get_yearly_filepath(dt)
    sheet_name = get_sheet_name(dt)

    if not filepath.exists():
        raise FileNotFoundError(
            f"연간 파일을 찾을 수 없습니다: {filepath}\n"
            f"파일을 INCIDENT_EXCEL_DIR({INCIDENT_EXCEL_DIR})에 복사해 주세요."
        )

    wb = openpyxl.load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트 '{sheet_name}'가 없습니다. 파일의 시트 목록: {wb.sheetnames}")

    ws = wb[sheet_name]

    # 중복 체크: 동일 제목이 이미 있으면 스킵
    if _is_duplicate(ws, record):
        print(f"  [skip] 이미 존재하는 레코드: {record.get('제목', '')[:30]}")
        return filepath

    next_id = _get_next_id(ws)
    record["ID"] = next_id
    target_row = _next_empty_row(ws)
    _write_row(ws, target_row, _build_row(record))
    _apply_style(ws, target_row)
    _update_count(ws, next_id)

    wb.save(filepath)
    print(f"  [incident_writer] {sheet_name} | ID {next_id} 추가 → {filepath.name}")
    return filepath


def process_incident(record: dict, entry: str, is_completion: bool = False, dt: datetime = None) -> Path:
    """
    서버명으로 현재 월 시트에서 활성(처리완료일시 없는) 행을 탐색.
    - 활성 행 있음: 조치내역에 entry 누적, 완료 시 처리완료일시 기록
    - 없음: 새 행 생성 후 entry 추가

    record에 "_server_name" 키로 서버명을 전달.
    """
    dt = dt or datetime.now()
    filepath = get_yearly_filepath(dt)
    sheet_name = get_sheet_name(dt)

    if not filepath.exists():
        raise FileNotFoundError(f"연간 파일을 찾을 수 없습니다: {filepath}")

    wb = openpyxl.load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트 '{sheet_name}'가 없습니다.")

    ws = wb[sheet_name]
    server_name = record.pop("_server_name", "") or ""
    active_row = _find_active_row(ws, server_name) if server_name else None

    if active_row:
        _append_entry(ws, active_row, entry)
        if is_completion:
            ws.cell(row=active_row, column=_COL["처리완료일시"]).value = datetime.now().strftime("%Y-%m-%d %H:%M")
        print(f"  [incident_writer] {sheet_name} | 행 {active_row} 업데이트 (완료={is_completion})")
    else:
        next_id = _get_next_id(ws)
        record["ID"] = next_id
        target_row = _next_empty_row(ws)
        _write_row(ws, target_row, _build_row(record))
        _append_entry(ws, target_row, entry)
        if is_completion:
            ws.cell(row=target_row, column=_COL["처리완료일시"]).value = datetime.now().strftime("%Y-%m-%d %H:%M")
        _apply_style(ws, target_row)
        _update_count(ws, next_id)
        print(f"  [incident_writer] {sheet_name} | ID {next_id} 새 행 생성 → {filepath.name}")

    wb.save(filepath)
    return filepath


def _find_active_row(ws, server_name: str) -> int | None:
    """제목 컬럼에 server_name 포함 + 처리완료일시 비어있는 첫 행 반환."""
    col_title = _COL["제목"]
    col_done = _COL["처리완료일시"]
    for row_num in range(3, ws.max_row + 1):
        title = ws.cell(row=row_num, column=col_title).value
        if not title or server_name.upper() not in str(title).upper():
            continue
        if not ws.cell(row=row_num, column=col_done).value:
            return row_num
    return None


def _append_entry(ws, row_num: int, entry: str):
    """조치내역 셀에 항목 누적 추가 (개행 구분)."""
    col = _COL["조치내역"]
    current = ws.cell(row=row_num, column=col).value or ""
    separator = "\n" if current else ""
    ws.cell(row=row_num, column=col).value = current + separator + entry


def append_bulk_incidents(records: list[dict], dt: datetime = None) -> Path:
    """여러 건을 한 번에 추가 (중복 자동 스킵)."""
    dt = dt or datetime.now()
    filepath = get_yearly_filepath(dt)
    sheet_name = get_sheet_name(dt)

    if not filepath.exists():
        raise FileNotFoundError(f"연간 파일을 찾을 수 없습니다: {filepath}")

    wb = openpyxl.load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트 '{sheet_name}'가 없습니다.")

    ws = wb[sheet_name]
    added = 0

    for record in records:
        if _is_duplicate(ws, record):
            continue
        next_id = _get_next_id(ws)
        record["ID"] = next_id
        target_row = _next_empty_row(ws)
        _write_row(ws, target_row, _build_row(record))
        _apply_style(ws, target_row)
        _update_count(ws, next_id)
        added += 1

    wb.save(filepath)
    print(f"  [incident_writer] {sheet_name} | {added}건 추가 → {filepath.name}")
    return filepath


def _next_empty_row(ws) -> int:
    """3행부터 ID 열(1열)이 비어 있는 첫 행 번호 반환."""
    for row_num in range(3, ws.max_row + 2):
        if ws.cell(row=row_num, column=1).value is None:
            return row_num
    return ws.max_row + 1


def _write_row(ws, row_num: int, values: list):
    for col_idx, val in enumerate(values, start=1):
        ws.cell(row=row_num, column=col_idx).value = val


def _get_next_id(ws) -> int:
    """마지막 데이터 행의 ID + 1 반환 (데이터는 3행부터)."""
    max_id = 0
    for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
        val = row[0]
        if isinstance(val, int) and val > max_id:
            max_id = val
    return max_id + 1


def _is_duplicate(ws, record: dict) -> bool:
    """제목이 동일한 행이 이미 있으면 True."""
    title = str(record.get("제목", "")).strip()
    if not title:
        return False
    for row in ws.iter_rows(min_row=3, min_col=12, max_col=12, values_only=True):
        if row[0] and str(row[0]).strip() == title:
            return True
    return False


def _build_row(record: dict) -> list:
    """COLUMNS 순서에 맞게 값 리스트 생성."""
    row = []
    for col in COLUMNS:
        val = record.get(col, "")
        # 처리시간(분): 숫자면 그대로, 문자면 변환 시도
        if col == "처리시간(분)" and isinstance(val, str) and val.strip("-").isdigit():
            val = int(val)
        # 건수: 숫자 추출 ("1건" → 1)
        if col == "건수" and isinstance(val, str):
            val = int("".join(filter(str.isdigit, val)) or 1)
        row.append(val)
    return row


def _apply_style(ws, row_num: int):
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border = _BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _update_count(ws, count: int):
    """1행 3열: 총 건수 업데이트."""
    ws.cell(row=1, column=3).value = count
